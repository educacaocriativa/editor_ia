"""
Leitor da planilha Taxonomia de Bloom (dados/bloom_taxonomia.xlsx).
Estrutura exata da planilha do usuário:
  Aba 1 — Descritivo do verbo principal: VERBO PRINCIPAL | DEFINIÇÃO TÉCNICA
  Aba 2 — Verbos auxiliares: VERBO | OBJETIVOS | PROCESSOS | RESULTANTES
"""
from pathlib import Path
from typing import Dict, List

try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

ARQUIVO_BLOOM = Path(__file__).parent.parent / "dados" / "bloom_taxonomia.xlsx"

# Estrutura completa extraída da planilha — usada como fallback
BLOOM_ESTRUTURA = {
    "Conhecimento": {
        "nivel": 1,
        "definicao": (
            "Habilidade do estudante em recordar, definir, reconhecer "
            "e identificar informações."
        ),
        "objetivos": [
            "especificar",
            "modos e meios para lidar com itens específicos",
            "fatos universais e abstrações num dado campo",
        ],
        "processos": [
            "definir", "reconhecer", "recitar", "identificar",
            "rotular", "compreender", "examinar", "mostrar",
            "coletar", "listar",
        ],
        "resultantes": ["rótulos", "nomes", "fatos", "definições", "conceitos"],
    },
    "Compreensão": {
        "nivel": 2,
        "definicao": (
            "Habilidade do estudante em demonstrar compreensão, traduzir, "
            "interpretar e extrapolar informações."
        ),
        "objetivos": ["tradução", "interpretação", "extrapolação"],
        "processos": [
            "traduzir", "interpretar", "explicar", "descrever",
            "resumir", "demonstrar",
        ],
        "resultantes": ["argumento", "explicação", "descrição", "resumo"],
    },
    "Aplicação": {
        "nivel": 3,
        "definicao": (
            "Habilidade do estudante em recolher e aplicar informações "
            "em situações específicas e concretas."
        ),
        "objetivos": [
            "uso de abstrações em situações específicas e concretas",
        ],
        "processos": [
            "aplicar", "solucionar", "experimentar", "demonstrar",
            "construir", "mostrar", "fazer", "ilustrar", "registrar",
        ],
        "resultantes": [
            "diagrama", "ilustração", "coleção", "mapa", "jogo",
            "quebra-cabeças", "modelo", "relato", "fotografia", "lição",
        ],
    },
    "Análise": {
        "nivel": 4,
        "definicao": (
            "Habilidade do estudante em estruturar informações, "
            "identificar elementos, relacionamentos e princípios."
        ),
        "objetivos": [
            "elementos",
            "relacionamentos",
            "princípios organizacionais",
        ],
        "processos": [
            "conectar", "relacionar", "diferenciar", "classificar",
            "arranjar", "estruturar", "agrupar", "interpretar",
            "organizar", "categorizar", "retirar", "comparar",
            "dissecar", "investigar",
        ],
        "resultantes": [
            "gráfico", "questionário", "categoria", "levantamento",
            "tabela", "delineamento", "diagrama", "conclusão",
            "lista", "plano", "resumo",
        ],
    },
    "Avaliação": {
        "nivel": 5,
        "definicao": (
            "Habilidade do estudante em fazer julgamentos com base "
            "em evidências internas e externas."
        ),
        "objetivos": [
            "julgamento em termos de evidência interna",
            "julgamento em termos de evidência externa",
        ],
        "processos": [
            "interpretar", "verificar", "julgar", "criticar",
            "decidir", "discutir", "disputar", "escolher",
        ],
        "resultantes": [
            "opinião", "julgamento", "recomendação", "veredito",
            "conclusão", "avaliação", "investigação", "editorial",
        ],
    },
    "Criação": {
        "nivel": 6,
        "definicao": (
            "Habilidade do estudante em estruturar informações para "
            "criar algo novo — comunicação inédita, plano ou conjunto "
            "de relacionamentos abstratos."
        ),
        "objetivos": [
            "comunicação inédita",
            "plano de operação",
            "conjunto de relacionamento abstratos",
        ],
        "processos": [
            "projetar", "reprojetar", "combinar", "consolidar",
            "agregar", "compor", "formular hipótese", "construir",
            "traduzir", "imaginar", "inventar", "criar",
            "inferir", "produzir", "predizer",
        ],
        "resultantes": [
            "poema", "projeto", "resumo", "fórmula", "invenção",
            "história", "solução", "máquina", "filme", "programa",
            "produto",
        ],
    },
}


def carregar_bloom_xlsx(caminho: str = None) -> Dict:
    """Carrega a planilha e retorna estrutura enriquecida."""
    path = Path(caminho) if caminho else ARQUIVO_BLOOM

    if not _OPENPYXL or not path.exists():
        return {"estrutura": BLOOM_ESTRUTURA, "disponivel": False, "fonte": "interno"}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        nomes = wb.sheetnames

        # Lê aba 1 — definições
        definicoes = {}
        if len(nomes) >= 1:
            ws1 = wb[nomes[0]]
            verbo_atual = None
            for row in ws1.iter_rows(min_row=2, values_only=True):
                v, d = (str(row[0]).strip() if row[0] else ""),                        (str(row[1]).strip() if len(row) > 1 and row[1] else "")
                if v:
                    verbo_atual = v.split(" ou ")[0].strip()
                if verbo_atual and d:
                    definicoes[verbo_atual] = d

        # Lê aba 2 — objetivos/processos/resultantes
        estrutura_xlsx = {}
        if len(nomes) >= 2:
            ws2 = wb[nomes[1]]
            verbo_atual = None
            for row in ws2.iter_rows(min_row=2, values_only=True):
                v = str(row[0]).strip() if row[0] else ""
                obj = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                proc = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                res = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                if v:
                    nome_chave = v.split(" ou ")[0].strip()
                    verbo_atual = nome_chave
                    if verbo_atual not in estrutura_xlsx:
                        estrutura_xlsx[verbo_atual] = {
                            "objetivos": [], "processos": [], "resultantes": []
                        }
                if verbo_atual:
                    if obj:
                        estrutura_xlsx[verbo_atual]["objetivos"].append(obj)
                    if proc:
                        estrutura_xlsx[verbo_atual]["processos"].append(proc)
                    if res:
                        estrutura_xlsx[verbo_atual]["resultantes"].append(res)

        # Mescla com estrutura interna (completa os níveis)
        estrutura_final = {}
        for nome_interno, dados_internos in BLOOM_ESTRUTURA.items():
            chave = nome_interno
            xlsx_dados = estrutura_xlsx.get(chave, {})
            estrutura_final[chave] = {
                "nivel": dados_internos["nivel"],
                "definicao": definicoes.get(chave, dados_internos["definicao"]),
                "objetivos": xlsx_dados.get("objetivos") or dados_internos["objetivos"],
                "processos": xlsx_dados.get("processos") or dados_internos["processos"],
                "resultantes": xlsx_dados.get("resultantes") or dados_internos["resultantes"],
            }

        return {
            "estrutura": estrutura_final,
            "disponivel": True,
            "fonte": str(path.name),
        }

    except Exception as e:
        return {
            "estrutura": BLOOM_ESTRUTURA,
            "disponivel": False,
            "fonte": "interno",
            "erro": str(e),
        }


def montar_contexto_bloom(caminho: str = None) -> str:
    """
    Retorna o conteúdo da planilha como bloco de texto para injeção no prompt.
    Formato detalhado com todos os 4 campos por nível.
    """
    dados = carregar_bloom_xlsx(caminho)
    estrutura = dados["estrutura"]

    linhas = [
        "=== TAXONOMIA DE BLOOM — BASE DE REFERÊNCIA ===",
        f"(Fonte: {dados.get('fonte', 'interno')})",
        "",
        "Para cada atividade, classifique usando EXATAMENTE estes campos:",
        "  verbo        → o nível principal (Conhecimento, Compreensão, etc.)",
        "  objetivos    → o que o verbo visa alcançar",
        "  processos    → o que o estudante faz cognitivamente",
        "  resultantes  → o que o estudante entrega como produto",
        "",
    ]

    for nome, dados_nivel in estrutura.items():
        n = dados_nivel["nivel"]
        linhas.append(f"━━ N{n} — {nome.upper()} ━━")
        linhas.append(f"Definição: {dados_nivel['definicao']}")
        linhas.append(
            f"Objetivos:   {' | '.join(dados_nivel['objetivos'][:3])}"
        )
        linhas.append(
            f"Processos:   {', '.join(dados_nivel['processos'][:8])}"
        )
        linhas.append(
            f"Resultantes: {', '.join(dados_nivel['resultantes'][:6])}"
        )
        linhas.append("")

    return "\n".join(linhas)


def obter_estrutura() -> Dict:
    """Retorna a estrutura completa de Bloom (planilha ou interna)."""
    return carregar_bloom_xlsx()["estrutura"]
