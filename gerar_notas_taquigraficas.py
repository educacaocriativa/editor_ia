"""
Gera a planilha de Notas Taquigráficas para revisão editorial.
Execute: python3 gerar_notas_taquigraficas.py
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────
# DADOS: (símbolo, nome, categoria, descrição, exemplo, observação)
# ──────────────────────────────────────────────────────────
NOTAS = [
    # ── SUPRESSÃO / INSERÇÃO ─────────────────────────────
    ("⌃ ou ^",        "Inserir",                "Supressão/Inserção",
     "Inserir texto, letra ou sinal no ponto indicado.",
     "O aluno^estudou muito.",
     "Escrever o texto a inserir acima ou na margem."),

    ("⌦ ou risca",    "Deletar / Suprimir",     "Supressão/Inserção",
     "Eliminar a palavra, letra ou trecho riscado.",
     "O aluno estudou [muito] bastante.",
     "Traçar linha sobre o trecho a excluir."),

    ("stet / sic",    "Manter / Não alterar",   "Supressão/Inserção",
     "'Stet' = manter o que foi marcado para exclusão. 'Sic' = registrar que o erro é do original.",
     "O aluno estudou [muito] bastante. stet",
     "Pontilhado sob o trecho indica 'stet'."),

    ("○",             "Ponto final",             "Pontuação",
     "Inserir ponto final no local indicado.",
     "O aluno estudou○",
     "Círculo pequeno no local exato."),

    (",",             "Vírgula",                 "Pontuação",
     "Inserir vírgula.",
     "Portanto, o resultado é correto,",
     "Acrescentar sinal na posição correta."),

    (";",             "Ponto e vírgula",         "Pontuação",
     "Inserir ponto e vírgula.",
     "Estudou muito; obteve boas notas.",
     ""),

    (":",             "Dois-pontos",             "Pontuação",
     "Inserir dois-pontos.",
     "Observe: a resposta está errada.",
     ""),

    ("?",             "Ponto de interrogação",   "Pontuação",
     "Inserir ponto de interrogação.",
     "Você entendeu?",
     ""),

    ("!",             "Ponto de exclamação",     "Pontuação",
     "Inserir ponto de exclamação.",
     "Que resultado excelente!",
     ""),

    ("…",             "Reticências",             "Pontuação",
     "Inserir reticências (3 pontos).",
     "Não sei… talvez.",
     "Usar apenas 3 pontos; não mais."),

    ("–",             "Travessão / Dash",        "Pontuação",
     "Inserir travessão (em diálogos ou aposto).",
     "– Bom dia – disse ela.",
     "Diferenciar de hífen."),

    ("-",             "Hífen",                   "Pontuação",
     "Inserir hífen (composição ou separação silábica).",
     "guarda-chuva / bem-vindo",
     "Não confundir com travessão."),

    ("« »  ou \" \"", "Aspas",                   "Pontuação",
     "Inserir aspas (citação, ironia, termo técnico).",
     "O conceito de \"equilíbrio\" é central.",
     "Usar aspas duplas em pt-BR."),

    ("( )",            "Parênteses",             "Pontuação",
     "Inserir parênteses.",
     "O resultado (ver tabela 1) confirma.",
     ""),

    ("[ ]",            "Colchetes",              "Pontuação",
     "Inserir colchetes (intervenção editorial).",
     "[N.E.: grifo nosso]",
     "Uso editorial para notas externas ao texto."),

    # ── TIPOGRAFIA / FORMATAÇÃO ──────────────────────────
    ("ital.",          "Itálico",                "Tipografia",
     "Formatar em itálico.",
     "ital. Homo sapiens",
     "Títulos de obras, estrangeirismos, ênfase."),

    ("negr. / bf",     "Negrito",                "Tipografia",
     "Formatar em negrito.",
     "negr. Conceito-chave",
     "Termos-chave, chamadas de atenção."),

    ("CAIXA / CAP",    "Caixa alta (maiúsculas)","Tipografia",
     "Converter para letras maiúsculas.",
     "CAP Brasil → BRASIL",
     "Siglas, títulos em destaque."),

    ("c/c",            "Caixa alta e baixa",     "Tipografia",
     "Inicial maiúscula, demais minúsculas.",
     "c/c BRASIL → Brasil",
     "Nomes próprios, títulos de seções."),

    ("rm / red.",      "Redondo (roman)",        "Tipografia",
     "Remover itálico/negrito; voltar para fonte regular.",
     "rm *palavra*  → palavra",
     "Desfaz formatação especial."),

    ("versalete",      "Versalete (small caps)", "Tipografia",
     "Formatar em versalete.",
     "versalete Nome → ɴᴏᴍᴇ",
     "Usado em epígrafes e créditos."),

    ("corpo ##",       "Tamanho de fonte",       "Tipografia",
     "Indicar tamanho de corpo de texto.",
     "corpo 12 → fonte 12pt",
     ""),

    # ── ESPAÇAMENTO / ESPACEJAMENTO ──────────────────────
    ("#",              "Espaço (inserir)",       "Espaçamento",
     "Inserir espaço entre palavras ou elementos.",
     "O#aluno → O aluno",
     ""),

    ("~  ou ⌢",       "Aproximar (tirar espaço)","Espaçamento",
     "Eliminar espaço desnecessário.",
     "O  aluno~→ O aluno",
     "Arco unindo os dois lados indica aproximar."),

    ("eq.",            "Equalizar espaço",       "Espaçamento",
     "Igualar espaçamento entre elementos.",
     "eq. [espaço irregular]",
     ""),

    ("ent.",           "Entrar (recuo/indent)",  "Espaçamento",
     "Adicionar recuo de parágrafo.",
     "ent. 1,25 cm",
     "Parágrafo sem recuo que deveria tê-lo."),

    ("s/ent.",         "Sem entrar (tirar recuo)","Espaçamento",
     "Remover recuo indevido.",
     "s/ent. [parágrafo com recuo errado]",
     ""),

    # ── ESTRUTURA / PARÁGRAFO ────────────────────────────
    ("¶",              "Parágrafo novo",         "Estrutura",
     "Iniciar novo parágrafo neste ponto.",
     "...fim da ideia.¶ Nova ideia...",
     "Símbolo pilcrow; inserir na quebra desejada."),

    ("no ¶",           "Juntar parágrafos",      "Estrutura",
     "Unir o parágrafo à linha anterior (sem quebra).",
     "no ¶ [entre dois parágrafos]",
     ""),

    ("run-in",         "Correr (sem quebra)",    "Estrutura",
     "Manter na mesma linha, sem quebra de parágrafo.",
     "run-in após o subtítulo",
     ""),

    ("tr",             "Transpor",               "Estrutura",
     "Inverter a ordem de palavras, frases ou elementos.",
     "tr O menino[correu rápido] → O menino rápido correu",
     "Indicar com setas ou arcos o que trocar."),

    ("mv",             "Mover",                  "Estrutura",
     "Deslocar bloco de texto para outro local.",
     "mv [§3 → após §5]",
     "Indicar origem e destino com setas."),

    ("fl.",            "Flush left (alinhar à esq.)","Estrutura",
     "Alinhar texto à margem esquerda.",
     "fl. [título centralizado errado]",
     ""),

    ("fr.",            "Flush right (alinhar à dir.)","Estrutura",
     "Alinhar texto à margem direita.",
     "fr. [crédito de imagem]",
     ""),

    ("ct.",            "Centralizar",            "Estrutura",
     "Centralizar o elemento.",
     "ct. [título de capítulo]",
     ""),

    ("just.",          "Justificar",             "Estrutura",
     "Alinhar texto nas duas margens.",
     "just. [corpo do texto]",
     ""),

    # ── INSTRUÇÃO GERAL / EDITORIAL ──────────────────────
    ("N.E.",           "Nota do editor",         "Instrução Editorial",
     "Observação inserida pelo editor, não pelo autor.",
     "[N.E.: ver também p. 34]",
     "Sempre entre colchetes e distinguida do texto."),

    ("N.A.",           "Nota do autor",          "Instrução Editorial",
     "Observação do próprio autor.",
     "[N.A.: ênfase adicionada]",
     ""),

    ("N.T.",           "Nota do tradutor",       "Instrução Editorial",
     "Observação inserida pelo tradutor.",
     "[N.T.: no original em inglês, 'resilience']",
     ""),

    ("cf.",            "Conferir / Verificar",   "Instrução Editorial",
     "Solicitar confirmação de dado, citação ou referência.",
     "cf. [autor/ano/página]",
     "O revisor deve verificar antes de publicar."),

    ("ver",            "Verificar / Checar",     "Instrução Editorial",
     "Indicar elemento a ser verificado (dado, link, imagem).",
     "ver [figura 3 — legenda]",
     ""),

    ("ok",             "Aprovado",               "Instrução Editorial",
     "Elemento revisado e aprovado sem alterações.",
     "ok [parágrafo]",
     "Confirma que está correto."),

    ("?",              "Dúvida editorial",       "Instrução Editorial",
     "Ponto duvidoso que requer decisão do autor/coordenador.",
     "? [termo técnico incomum]",
     "Inserir na margem com comentário."),

    ("att.",           "Atenção",                "Instrução Editorial",
     "Marcar ponto crítico ou inconsistência grave.",
     "att. [contradição com p. 12]",
     ""),

    ("rep.",           "Repetição",              "Instrução Editorial",
     "Indicar palavra ou trecho repetido desnecessariamente.",
     "rep. [palavra usada 3x no mesmo parágrafo]",
     ""),

    ("fig.",           "Figura / Imagem",        "Instrução Editorial",
     "Indicar inserção ou ajuste de figura.",
     "fig. [inserir gráfico aqui]",
     "Especificar número e legenda."),

    ("tab.",           "Tabela",                 "Instrução Editorial",
     "Indicar inserção ou ajuste de tabela.",
     "tab. [tabela 2 — ajustar alinhamento]",
     ""),

    ("quad.",          "Quadro",                 "Instrução Editorial",
     "Indicar inserção ou ajuste de quadro.",
     "quad. [quadro comparativo]",
     ""),

    ("ref.",           "Referência bibliográfica","Instrução Editorial",
     "Verificar ou completar referência.",
     "ref. [autor, ano, título incompleto]",
     "ABNT NBR 6023."),

    ("ibid.",          "Ibidem (mesma obra)",    "Instrução Editorial",
     "Remeter à mesma obra citada anteriormente.",
     "ibid., p. 45.",
     ""),

    ("op. cit.",       "Opere citato",           "Instrução Editorial",
     "Remeter a obra já citada anteriormente (não a imediata).",
     "SILVA, op. cit., p. 30.",
     ""),

    # ── ESPECÍFICAS PARA MATERIAL DIDÁTICO ───────────────
    ("ativ.",          "Atividade",              "Material Didático",
     "Marcar ou ajustar enunciado de atividade.",
     "ativ. [enunciado pouco claro]",
     "Verificar alinhamento com habilidade BNCC."),

    ("BNCC",           "Habilidade BNCC",        "Material Didático",
     "Indicar ou verificar habilidade da Base Nacional.",
     "BNCC EF05CI01",
     "Código no padrão EF/EM + ano + área + número."),

    ("obj.",           "Objetivo de aprendizagem","Material Didático",
     "Verificar/inserir objetivo de aprendizagem.",
     "obj. [não declarado no início da unidade]",
     ""),

    ("cmd.",           "Comando de atividade",   "Material Didático",
     "Revisar o comando (verbo de ação) da atividade.",
     "cmd. [verbo vago: 'fale sobre' → 'descreva']",
     "Usar verbos operacionais de Bloom."),

    ("ling.",          "Adequação linguística",  "Material Didático",
     "Linguagem inadequada para a faixa etária.",
     "ling. [vocabulário acima do nível EF1]",
     "Ajustar ao perfil etário do perfil configurado."),

    ("incl.",          "Inclusão / Acessibilidade","Material Didático",
     "Verificar linguagem inclusiva e acessível.",
     "incl. [termo excludente]",
     "Seguir diretrizes de linguagem inclusiva."),

    ("imagem ok",      "Imagem pedagogicamente adequada","Material Didático",
     "Confirmar que a imagem é adequada ao conteúdo e faixa etária.",
     "imagem ok [fig. 2]",
     ""),

    ("gabarito",       "Verificar gabarito",     "Material Didático",
     "Conferir resposta esperada da atividade.",
     "gabarito [questão 3 — resposta diverge do texto]",
     ""),
]

CABECALHO = [
    "Símbolo / Marca",
    "Nome",
    "Categoria",
    "Descrição / Significado",
    "Exemplo de Uso",
    "Observações",
]

# ── PALETA DE CORES POR CATEGORIA ────────────────────────
COR_CATEGORIA = {
    "Supressão/Inserção":   "FFF2CC",   # amarelo suave
    "Pontuação":            "E2EFDA",   # verde suave
    "Tipografia":           "DDEEFF",   # azul suave
    "Espaçamento":          "FCE4D6",   # laranja suave
    "Estrutura":            "EAD1DC",   # rosa suave
    "Instrução Editorial":  "D9D9D9",   # cinza
    "Material Didático":    "D0E4F7",   # azul médio
}
COR_CABECALHO  = "2F5597"  # azul escuro
FONTE_CABECALHO = "FFFFFF"


def estilo_borda():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def criar_planilha(caminho: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas Taquigráficas"

    # ── Cabeçalho ────────────────────────────────────────
    ws.append(CABECALHO)
    for col_idx, _ in enumerate(CABECALHO, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color=FONTE_CABECALHO, size=11)
        cell.fill      = PatternFill("solid", fgColor=COR_CABECALHO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = estilo_borda()
    ws.row_dimensions[1].height = 28

    # ── Linhas de dados ───────────────────────────────────
    for row_idx, (simbolo, nome, cat, desc, exemplo, obs) in enumerate(NOTAS, start=2):
        ws.append([simbolo, nome, cat, desc, exemplo, obs])
        cor = COR_CATEGORIA.get(cat, "FFFFFF")
        fill = PatternFill("solid", fgColor=cor)
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = estilo_borda()
            if col_idx == 1:  # símbolo em negrito
                cell.font = Font(bold=True, size=11)
            elif col_idx == 3:  # categoria centralizada
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 38

    # ── Larguras de coluna ────────────────────────────────
    larguras = [18, 26, 22, 52, 48, 42]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Congelar painel no cabeçalho ──────────────────────
    ws.freeze_panes = "A2"

    # ── Filtro automático ─────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Aba de legenda de categorias ─────────────────────
    ws2 = wb.create_sheet("Legenda de Categorias")
    ws2.append(["Categoria", "Cor", "Uso Principal"])
    ws2["A1"].font = Font(bold=True, color=FONTE_CABECALHO)
    ws2["A1"].fill = PatternFill("solid", fgColor=COR_CABECALHO)
    ws2["B1"].font = Font(bold=True, color=FONTE_CABECALHO)
    ws2["B1"].fill = PatternFill("solid", fgColor=COR_CABECALHO)
    ws2["C1"].font = Font(bold=True, color=FONTE_CABECALHO)
    ws2["C1"].fill = PatternFill("solid", fgColor=COR_CABECALHO)

    descricoes_cat = {
        "Supressão/Inserção":   "Inserir ou remover texto, letras e sinais",
        "Pontuação":            "Todos os sinais de pontuação",
        "Tipografia":           "Formatação de fonte (itálico, negrito, caixa)",
        "Espaçamento":          "Controle de espaços, recuos e alinhamento",
        "Estrutura":            "Parágrafos, transposição, movimento de blocos",
        "Instrução Editorial":  "Notas, dúvidas, aprovações e chamadas do editor",
        "Material Didático":    "Específicas para revisão de livros didáticos (BNCC)",
    }
    for r, (cat, cor) in enumerate(COR_CATEGORIA.items(), start=2):
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2, value="████")
        ws2.cell(row=r, column=2).fill = PatternFill("solid", fgColor=cor)
        ws2.cell(row=r, column=3, value=descricoes_cat.get(cat, ""))
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = estilo_borda()
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 55

    wb.save(caminho)
    print(f"✅  Planilha salva em: {caminho}")
    print(f"   {len(NOTAS)} notas taquigráficas em {len(COR_CATEGORIA)} categorias.")


if __name__ == "__main__":
    import os
    destino = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "dados",
        "Notas_Taquigraficas_Editorial.xlsx"
    )
    os.makedirs(os.path.dirname(destino), exist_ok=True)
    criar_planilha(destino)
